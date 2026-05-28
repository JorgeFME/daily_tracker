"""
Generador de reporte Excel de ausencias del equipo.

Layout:
  - Fila 1: Título del reporte (mes/año)
  - Fila 2: Encabezados fijos + abreviatura día semana por día del mes
  - Fila 3: Números de días del mes
  - Fila 4+: Una por registro de ausencia, con marcas de color+texto en rango

Columnas fijas:
  A: Número de Empleado
  B: Nombre y Apellidos
  C: Inicio de Periodo
  D: Fin de Periodo
  E: Días de Ausencia (días calendario dentro del mes)
  F…: Días 1-N del mes
"""
import io
import calendar
import re
import unicodedata
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta ───────────────────────────────────────────────────────────────────
C_DARK  = "1E293B"
C_MID   = "334155"
C_WHITE = "FFFFFF"
C_GRAY  = "F1F5F9"
C_LIGHT = "F8FAFC"

# Tonos suaves para encabezado de calendario (evita colores agresivos).
C_CAL_WEEKDAY_BG   = "DBEAFE"
C_CAL_WEEKDAY_FG   = "1E3A8A"
C_CAL_WEEKEND_BG   = "FEE2E2"
C_CAL_WEEKEND_FG   = "9F1239"
C_CAL_DAYNUM_WK_BG = "FFF1F2"

# (fondo, texto, abreviatura) — colores alineados con calendario.html
TIPO_COLOR: dict[str, tuple[str, str, str]] = {
    "VACACIONES":  ("E0F2FE", "0369A1", "VA"),
    "INCAPACIDAD": ("FCE7F3", "9D174D", "INC"),
    "DIA_LIBRE":   ("DCFCE7", "14532D", "DL"),
    "PERMISO":     ("FEF9C3", "854D0E", "PER"),
    "OTRO":        ("F3E8FF", "6B21A8", "OTR"),
}
TIPO_DEFAULT = ("F1F5F9", "475569", "AU")

TIPOS_VALIDOS = set(TIPO_COLOR.keys())

THIN   = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
C_ALN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ALN  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

DIAS_ES  = ["L", "M", "M", "J", "V", "S", "D"]  # weekday() 0=Lun
MESES_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

# Columna (1-indexed) a partir de la cual empiezan los días del mes
_FIRST_DAY_COL = 6  # F


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color, start_color=color)


def _font(bold: bool = False, color: str = C_DARK, size: int = 9,
          italic: bool = False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _col_dia(dia: int) -> int:
    """Columna openpyxl (1-indexed) para el día N del mes."""
    return _FIRST_DAY_COL + dia - 1


def _parse_ymd(val) -> str | None:
    """Extrae 'YYYY-MM-DD' de distintos tipos de valor que retorna HANA."""
    if val is None:
        return None
    text = str(val).strip()
    m = re.search(r"\d{4}-\d{2}-\d{2}", text)
    return m.group(0) if m else None


def _fmt_ddmmyyyy(ymd: str) -> str:
    """Convierte una fecha YYYY-MM-DD a DD-MM-YYYY para mostrar en Excel."""
    try:
        d = date.fromisoformat(ymd)
        return d.strftime("%d-%m-%Y")
    except Exception:
        return ymd


def _contar_dias_habiles(inicio: date, fin: date) -> int:
    """Cuenta solo lunes-viernes en un rango inclusivo."""
    if fin < inicio:
        return 0
    dias = 0
    actual = inicio
    while actual <= fin:
        if actual.weekday() < 5:
            dias += 1
        actual = date.fromordinal(actual.toordinal() + 1)
    return dias


def _normalizar_nombre(texto: str) -> str:
    """Normaliza nombre para comparar sin acentos ni mayusculas."""
    limpio = unicodedata.normalize("NFKD", str(texto or ""))
    limpio = "".join(ch for ch in limpio if not unicodedata.combining(ch))
    limpio = re.sub(r"\s+", " ", limpio).strip().lower()
    return limpio


def _numero_empleado_por_nombre(nombre: str) -> str:
    """Regresa numero de empleado hardcodeado; si no existe, retorna NA."""
    nombre_norm = _normalizar_nombre(nombre)

    # Hardcode solicitado por negocio mientras no exista campo en BD.
    catalogo = {
        "analia": "10061763",
        "ana": "10061763",
        "oscar": "10054837",
        "bernardo": "10062038",
        "jorge": "10064030",
        "mario": "10062446",
    }

    for alias, numero in catalogo.items():
        if re.search(rf"\b{re.escape(alias)}\b", nombre_norm):
            return numero

    return "NA"


def generar_reporte_ausencias(
    anio: int,
    mes: int,
    empleados: list,
    ausencias: list,
    tipo_filtro: str = "",
) -> bytes:
    """
    Genera el reporte Excel de ausencias como bytes XLSX.

    Args:
        anio:         Año del reporte (2020-2100).
        mes:          Mes del reporte (1-12).
        empleados:    Lista de dicts con keys ID, NOMBRE_COMPLETO.
        ausencias:    Resultado de obtener_ausencias_mes(); puede incluir todos los tipos.
        tipo_filtro:  Si se indica (ej. 'VACACIONES'), filtra solo ese tipo.

    Returns:
        bytes del archivo XLSX listo para enviar con send_file.
    """
    # ── Validaciones de parámetros ────────────────────────────────────────────
    if not (2020 <= anio <= 2100):
        raise ValueError("Año debe estar entre 2020 y 2100.")
    if not (1 <= mes <= 12):
        raise ValueError("Mes debe estar entre 1 y 12.")
    tipo_filtro = (tipo_filtro or "").strip().upper()
    if tipo_filtro and tipo_filtro not in TIPOS_VALIDOS:
        raise ValueError(f"Tipo de ausencia inválido: {tipo_filtro!r}.")

    # ── Normalizar y filtrar ausencias ────────────────────────────────────────
    rows: list[dict] = []
    for a in ausencias:
        fi = _parse_ymd(a.get("FECHA_INICIO") or a.get("fecha_inicio"))
        ff = _parse_ymd(a.get("FECHA_FIN")    or a.get("fecha_fin"))
        if not fi or not ff:
            continue
        tipo = (a.get("TIPO") or a.get("tipo") or "OTRO").strip().upper()
        if tipo_filtro and tipo != tipo_filtro:
            continue
        rows.append({
            "id_usuario":   str(a.get("ID_USUARIO") or a.get("id_usuario") or ""),
            "usuario":      str(a.get("USUARIO")    or a.get("usuario")    or ""),
            "fecha_inicio": fi,
            "fecha_fin":    ff,
            "tipo":         tipo,
        })

    rows.sort(key=lambda r: (r["fecha_inicio"], r["usuario"]))

    # ── Metadatos del mes ─────────────────────────────────────────────────────
    num_dias   = calendar.monthrange(anio, mes)[1]
    primer_dia = date(anio, mes, 1)
    ultimo_dia = date(anio, mes, num_dias)

    # La columna A se mantiene como NA por requerimiento funcional.

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESES_ES[mes - 1][:3]}-{str(anio)[-2:]}"

    # Anchos de columnas fijas
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    for dia in range(1, num_dias + 1):
        ws.column_dimensions[get_column_letter(_col_dia(dia))].width = 3.5

    total_cols = _col_dia(num_dias)

    # ─────────────────────────────────────────────────────────────────────────
    # FILA 1 — Título
    # ─────────────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    tipo_label = ""
    if tipo_filtro and tipo_filtro in TIPO_COLOR:
        nombre_tipo = tipo_filtro.replace("_", " ").title()
        tipo_label  = f"  ·  {nombre_tipo}"

    titulo = ws.cell(row=1, column=1)
    titulo.value     = f"Reporte de Ausencias{tipo_label}  ·  {MESES_ES[mes - 1]} {anio}"
    titulo.font      = _font(bold=True, color=C_WHITE, size=11)
    titulo.fill      = _fill(C_DARK)
    titulo.alignment = L_ALN
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    # Rellenar celdas mergeadas para que tengan el mismo fondo
    for col in range(2, total_cols + 1):
        ws.cell(row=1, column=col).fill = _fill(C_DARK)

    # ─────────────────────────────────────────────────────────────────────────
    # FILA 2 — Encabezados columnas fijas + día de semana
    # ─────────────────────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 28

    FIXED_HEADERS = [
        (1, "Número de\nEmpleado"),
        (2, "Nombre y Apellidos"),
        (3, "Inicio de\nPeriodo"),
        (4, "Fin de\nPeriodo"),
        (5, "Días de\nVacaciones"),
    ]
    for col, label in FIXED_HEADERS:
        c = ws.cell(row=2, column=col)
        c.value     = label
        c.font      = _font(bold=True, color=C_WHITE, size=8)
        c.fill      = _fill(C_MID)
        c.alignment = C_ALN
        c.border    = BORDER

    for dia in range(1, num_dias + 1):
        dow  = date(anio, mes, dia).weekday()   # 0=Lun, 6=Dom
        col  = _col_dia(dia)
        fondo = C_CAL_WEEKDAY_BG if dow < 5 else C_CAL_WEEKEND_BG
        texto = C_CAL_WEEKDAY_FG if dow < 5 else C_CAL_WEEKEND_FG
        c = ws.cell(row=2, column=col)
        c.value     = DIAS_ES[dow]
        c.font      = _font(bold=True, color=texto, size=7)
        c.fill      = _fill(fondo)
        c.alignment = C_ALN
        c.border    = BORDER

    # ─────────────────────────────────────────────────────────────────────────
    # FILA 3 — Números de día
    # ─────────────────────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 14

    for col in range(1, 6):
        ws.cell(row=3, column=col).fill   = _fill(C_GRAY)
        ws.cell(row=3, column=col).border = BORDER

    for dia in range(1, num_dias + 1):
        dow = date(anio, mes, dia).weekday()
        col = _col_dia(dia)
        c = ws.cell(row=3, column=col)
        c.value     = dia
        c.font      = _font(bold=True, color=C_CAL_WEEKEND_FG if dow >= 5 else C_DARK, size=7)
        c.fill      = _fill(C_CAL_DAYNUM_WK_BG if dow >= 5 else C_LIGHT)
        c.alignment = C_ALN
        c.border    = BORDER

    # Congelar las 3 filas de encabezado y las 5 columnas fijas
    ws.freeze_panes = "F4"

    # ─────────────────────────────────────────────────────────────────────────
    # FILAS DE DATOS
    # ─────────────────────────────────────────────────────────────────────────
    DATA_START = 4

    if not rows:
        c = ws.cell(row=DATA_START, column=1)
        c.value     = "Sin ausencias registradas para el período seleccionado."
        c.font      = _font(italic=True, color="94A3B8", size=9)
        c.alignment = L_ALN
        ws.merge_cells(start_row=DATA_START, start_column=1,
                       end_row=DATA_START,   end_column=total_cols)
    else:
        for r_idx, row in enumerate(rows):
            excel_row = DATA_START + r_idx
            ws.row_dimensions[excel_row].height = 16

            tipo                      = row["tipo"]
            bg_tipo, fg_tipo, short   = TIPO_COLOR.get(tipo, TIPO_DEFAULT)
            row_bg                    = C_WHITE if r_idx % 2 == 0 else C_LIGHT

            # A — Numero de Empleado (hardcodeado por nombre)
            _set(ws, excel_row, 1, _numero_empleado_por_nombre(row["usuario"]), row_bg, bold=False, aln=C_ALN)

            # B — Nombre y Apellidos
            _set(ws, excel_row, 2, row["usuario"], row_bg, bold=True, aln=L_ALN)

            # C — Inicio de Periodo
            _set(ws, excel_row, 3, _fmt_ddmmyyyy(row["fecha_inicio"]), row_bg, bold=False, aln=C_ALN)

            # D — Fin de Periodo
            _set(ws, excel_row, 4, _fmt_ddmmyyyy(row["fecha_fin"]), row_bg, bold=False, aln=C_ALN)

            # E — Días de ausencia (solo hábiles, recortados al mes)
            fi_d   = date.fromisoformat(row["fecha_inicio"])
            ff_d   = date.fromisoformat(row["fecha_fin"])
            fi_mes = max(fi_d, primer_dia)
            ff_mes = min(ff_d, ultimo_dia)
            dias_en_mes = _contar_dias_habiles(fi_mes, ff_mes)
            _set(ws, excel_row, 5, dias_en_mes, row_bg, bold=True, aln=C_ALN)

            # F … — Marcas por día
            for dia in range(1, num_dias + 1):
                fecha_dia = date(anio, mes, dia)
                col       = _col_dia(dia)
                dow       = fecha_dia.weekday()

                if fi_d <= fecha_dia <= ff_d and dow < 5:
                    # Día cubierto por la ausencia
                    c = ws.cell(row=excel_row, column=col)
                    c.value     = short
                    c.font      = _font(bold=True, color=fg_tipo, size=7)
                    c.fill      = _fill(bg_tipo)
                    c.alignment = C_ALN
                    c.border    = BORDER
                else:
                    # Día no cubierto; fin de semana con fondo suave diferente
                    fondo = "FEF2F2" if dow >= 5 else row_bg
                    c = ws.cell(row=excel_row, column=col)
                    c.fill   = _fill(fondo)
                    c.border = BORDER

    # ── Guardar ───────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helper interno ────────────────────────────────────────────────────────────
def _set(ws, row: int, col: int, value, bg: str,
         bold: bool = False, aln: Alignment = None) -> None:
    """Escribe una celda de datos con estilo estándar."""
    c = ws.cell(row=row, column=col)
    c.value     = value
    c.font      = _font(bold=bold, size=8)
    c.fill      = _fill(bg)
    c.alignment = aln or C_ALN
    c.border    = BORDER
