"""
Generador de reporte Excel filtrado desde Registros.

- Hoja 1: Resumen ejecutivo
- Hoja 2: Detalle de actividades
- Evidencia resumida por tipo
"""
import io
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta ──────────────────────────────────────────────────────────────────
C_DARK = "1E293B"
C_MID = "334155"
C_BLUE = "2563EB"
C_LBLUE = "DBEAFE"
C_GRAY = "F8FAFC"
C_WHITE = "FFFFFF"
C_ALTA = "FEE2E2"
C_MEDIA = "FEF9C3"
C_BAJA = "DCFCE7"
C_LINKED = "F0F4FF"
C_PARENT = "EDE9FE"
C_WARN = "FFF7ED"
C_WARN_BORDER = "FDBA74"
C_SUCCESS = "DCFCE7"

ESTATUS_BG = {
    "en análisis": "EFF6FF",
    "en proceso": "DBEAFE",
    "en espera de info del cliente": "FEF3C7",
    "completado": "DCFCE7",
    "cancelado": "F1F5F9",
}

THIN = Side(style="thin", color="CBD5E1")
WARN_SIDE = Side(style="thin", color=C_WARN_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WARN_BORDER_STYLE = Border(left=WARN_SIDE, right=WARN_SIDE, top=WARN_SIDE, bottom=WARN_SIDE)
C_ALN = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ALN = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ALN = C_ALN


def _fill(color):
    return PatternFill("solid", fgColor=color, start_color=color)


def _font(bold=False, color=C_DARK, size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _prioridad(value):
    return {1: "Alta", 2: "Media", 3: "Baja"}.get(int(value), "—") if value is not None else "—"


def _fecha(value):
    if value is None:
        return "—"
    text = str(value).strip()
    return text[:10] if len(text) >= 10 else text


def _safe(value):
    text = str(value).strip() if value is not None else ""
    return text or "—"


def _es_actividad_rapida(row):
    return str(row.get("NOMBRE_ACTIVIDAD") or "").startswith("⚡")


def _tipos_evidencia_texto(evidencias):
    tipos = []
    for evidencia in evidencias:
        tipo = str(evidencia.get("TIPO") or "").strip()
        if tipo and tipo not in tipos:
            tipos.append(tipo)
    return f"Sí · {' / '.join(tipos)}" if tipos else "Sin evidencia"


def _linked_text(row):
    num_hijas = int(row.get("NUM_HIJAS") or 0)
    nombres_hijas = row.get("NOMBRES_HIJAS") or ""
    padre_id = row.get("ID_ACTIVIDAD_PADRE")
    nombre_padre = row.get("NOMBRE_ACTIVIDAD_PADRE") or f"ID {padre_id}"

    partes = []
    if padre_id:
        partes.append("Derivada de:")
        partes.append(f"  - {nombre_padre}")

    if num_hijas > 0:
        if padre_id:
            partes.append("")
        etiqueta = "actividad derivada" if num_hijas == 1 else "actividades derivadas"
        partes.append(f"{num_hijas} {etiqueta}:")
        if nombres_hijas:
            for nombre_hija in nombres_hijas.split(" | "):
                partes.append(f"  - {nombre_hija.strip()}")

    return "\n".join(partes) if partes else "—"


def _build_summary(actividades, evidencias_por_actividad):
    total_actividades = len(actividades)
    total_horas = sum(float(item.get("HORAS_TOTALES") or 0) for item in actividades)
    actividades_con_evidencia = sum(
        1 for item in actividades if evidencias_por_actividad.get(item.get("ID"))
    )
    actividades_rapidas = sum(1 for item in actividades if _es_actividad_rapida(item))
    actividades_rapidas_historicas = sum(
        1 for item in actividades if int(item.get("ES_ACTIVIDAD_RAPIDA_HISTORICA") or 0) == 1
    )

    estatus_counter = Counter()
    estatus_horas = Counter()
    prioridad_counter = Counter()
    for item in actividades:
        estatus = _safe(item.get("ESTATUS"))
        horas = float(item.get("HORAS_TOTALES") or 0)
        estatus_counter[estatus] += 1
        estatus_horas[estatus] += horas

        prioridad = _prioridad(item.get("PRIORIDAD"))
        if prioridad != "—":
            prioridad_counter[prioridad] += 1

    tipo_counter = Counter()
    tipo_horas = Counter()
    for item in actividades:
        tipo = _safe(item.get("TIPO"))
        horas_item = float(item.get("HORAS_TOTALES") or 0)
        tipo_counter[tipo] += 1
        tipo_horas[tipo] += horas_item

    return {
        "total_actividades": total_actividades,
        "total_horas": total_horas,
        "actividades_con_evidencia": actividades_con_evidencia,
        "actividades_rapidas": actividades_rapidas,
        "actividades_rapidas_historicas": actividades_rapidas_historicas,
        "estatus_counter": estatus_counter,
        "estatus_horas": estatus_horas,
        "prioridad_counter": prioridad_counter,
        "tipo_counter": tipo_counter,
        "tipo_horas": tipo_horas,
    }


def _export_scope(export_context=None):
    export_context = export_context or {}
    return export_context.get("scope_label") or "Todas las actividades del proyecto exportado"


def _export_scope_short(export_context=None):
    export_context = export_context or {}
    return export_context.get("scope_label_short") or "Detalle cronológico por actividad"


def _render_resumen(ws, proyecto_nombre, generado_en, actividades, evidencias_por_actividad, export_context=None):
    summary = _build_summary(actividades, evidencias_por_actividad)

    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = "1E293B"
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = f"Resumen ejecutivo · {proyecto_nombre}"
    title.font = _font(bold=True, color=C_WHITE, size=16)
    title.fill = _fill(C_DARK)
    title.alignment = C_ALN
    title.border = BORDER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    subtitle = ws["A2"]
    subtitle.value = (
        "Reporte orientado a cliente, PM y gerencia. Combina resumen ejecutivo, "
        "trazabilidad de exportación y detalle cronológico por actividad."
    )
    subtitle.font = _font(color=C_MID, size=9)
    subtitle.fill = _fill(C_GRAY)
    subtitle.alignment = L_ALN
    subtitle.border = BORDER
    ws.row_dimensions[2].height = 34

    ws.merge_cells("A4:F4")
    section = ws["A4"]
    section.value = "Contexto de exportación"
    section.font = _font(bold=True, color=C_WHITE, size=10)
    section.fill = _fill(C_MID)
    section.alignment = L_ALN
    section.border = BORDER

    context_rows = [
        ("Proyecto", proyecto_nombre),
        ("Generado el", generado_en),
        ("Cobertura", _export_scope(export_context)),
        ("Criterio de evidencia", "La evidencia se resume por tipo; no se incrustan archivos"),
    ]
    for index, (label, value) in enumerate(context_rows, start=5):
        left = ws.cell(row=index, column=1, value=label)
        right = ws.cell(row=index, column=2, value=value)
        for cell in (left, right):
            cell.border = BORDER
            cell.alignment = L_ALN
            cell.fill = _fill(C_WHITE if index % 2 else C_GRAY)
        left.font = _font(bold=True, color=C_MID)
        right.font = _font()
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=6)

    ws.merge_cells("A10:C10")
    summary_header = ws["A10"]
    summary_header.value = "Resumen general"
    summary_header.font = _font(bold=True, color=C_WHITE, size=10)
    summary_header.fill = _fill(C_BLUE)
    summary_header.alignment = L_ALN
    summary_header.border = BORDER

    general_rows = [
        ("Actividades totales", summary["total_actividades"]),
        ("Horas totales", round(summary["total_horas"], 2)),
        ("Actividades con evidencia", summary["actividades_con_evidencia"]),
        ("Actividades rápidas", summary["actividades_rapidas"]),
        ("Rápidas históricas con datos parciales", summary["actividades_rapidas_historicas"]),
    ]
    for index, (label, value) in enumerate(general_rows, start=11):
        left = ws.cell(row=index, column=1, value=label)
        right = ws.cell(row=index, column=2, value=value)
        for cell in (left, right):
            cell.border = BORDER
            cell.alignment = C_ALN if cell.column == 2 else L_ALN
            cell.fill = _fill(C_WHITE if index % 2 else C_GRAY)
        left.font = _font(bold=True, color=C_MID)
        right.font = _font(bold=True, color=C_BLUE)
        if label == "Horas totales":
            right.number_format = "#,##0.0"
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=3)

    ws.merge_cells("D10:F10")
    status_header = ws["D10"]
    status_header.value = "Distribución por estatus"
    status_header.font = _font(bold=True, color=C_WHITE, size=10)
    status_header.fill = _fill(C_BLUE)
    status_header.alignment = L_ALN
    status_header.border = BORDER

    ws["D11"] = "Estatus"
    ws["E11"] = "Actividades"
    ws["F11"] = "Horas"
    for cell in (ws["D11"], ws["E11"], ws["F11"]):
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_MID)
        cell.alignment = C_ALN
        cell.border = BORDER

    status_start_row = 12
    for offset, estatus in enumerate(sorted(summary["estatus_counter"].keys()), start=0):
        row = status_start_row + offset
        ws.cell(row=row, column=4, value=estatus)
        ws.cell(row=row, column=5, value=summary["estatus_counter"][estatus])
        ws.cell(row=row, column=6, value=round(summary["estatus_horas"][estatus], 2))
        for column in range(4, 7):
            cell = ws.cell(row=row, column=column)
            cell.border = BORDER
            cell.fill = _fill(C_WHITE if row % 2 else C_GRAY)
            cell.alignment = C_ALN if column in (5, 6) else L_ALN
        ws.cell(row=row, column=6).number_format = "#,##0.0"

    prioridad_section_row = max(status_start_row + max(len(summary["estatus_counter"]), 1) + 2, 18)
    ws.merge_cells(start_row=prioridad_section_row, start_column=1, end_row=prioridad_section_row, end_column=2)
    prioridad_header = ws.cell(row=prioridad_section_row, column=1, value="Distribución por prioridad")
    prioridad_header.font = _font(bold=True, color=C_WHITE, size=10)
    prioridad_header.fill = _fill(C_BLUE)
    prioridad_header.alignment = L_ALN
    prioridad_header.border = BORDER

    prioridad_labels = ("Prioridad", "Actividades")
    for offset, label in enumerate(prioridad_labels, start=0):
        cell = ws.cell(row=prioridad_section_row + 1, column=1 + offset, value=label)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_MID)
        cell.alignment = C_ALN
        cell.border = BORDER

    for offset, prioridad in enumerate(("Alta", "Media", "Baja"), start=0):
        row = prioridad_section_row + 2 + offset
        left = ws.cell(row=row, column=1, value=prioridad)
        right = ws.cell(row=row, column=2, value=summary["prioridad_counter"].get(prioridad, 0))
        left.border = right.border = BORDER
        left.alignment = L_ALN
        right.alignment = C_ALN
        left.fill = right.fill = _fill(C_WHITE if row % 2 else C_GRAY)
        left.font = _font(bold=True)
        right.font = _font(bold=True, color=C_BLUE)

    alert_row = prioridad_section_row + 6
    ws.merge_cells(start_row=alert_row, start_column=1, end_row=alert_row, end_column=6)
    alert_header = ws.cell(row=alert_row, column=1, value="Alertas y notas")
    alert_header.font = _font(bold=True, color=C_WHITE, size=10)
    alert_header.fill = _fill(C_MID)
    alert_header.alignment = L_ALN
    alert_header.border = BORDER

    alert_text = (
        "Se detectaron actividades rápidas históricas con datos parciales. "
        "En la hoja 'Actividades' se marcan explícitamente para evitar interpretaciones incompletas."
        if summary["actividades_rapidas_historicas"] > 0
        else "No se detectaron alertas de integridad en las actividades exportadas."
    )
    ws.merge_cells(start_row=alert_row + 1, start_column=1, end_row=alert_row + 2, end_column=6)
    alert_note = ws.cell(row=alert_row + 1, column=1, value=alert_text)
    alert_note.font = _font(color=C_DARK, italic=summary["actividades_rapidas_historicas"] == 0)
    alert_note.fill = _fill(C_WARN if summary["actividades_rapidas_historicas"] > 0 else C_SUCCESS)
    alert_note.border = WARN_BORDER_STYLE if summary["actividades_rapidas_historicas"] > 0 else BORDER
    alert_note.alignment = L_ALN

    tipo_section_row = alert_row + 4
    ws.merge_cells(start_row=tipo_section_row, start_column=1, end_row=tipo_section_row, end_column=3)
    tipo_hdr = ws.cell(row=tipo_section_row, column=1, value="Distribución por tipo")
    tipo_hdr.font = _font(bold=True, color=C_WHITE, size=10)
    tipo_hdr.fill = _fill(C_BLUE)
    tipo_hdr.alignment = L_ALN
    tipo_hdr.border = BORDER

    for col_i, label in enumerate(("Tipo", "Actividades", "Horas"), start=1):
        cell = ws.cell(row=tipo_section_row + 1, column=col_i, value=label)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_MID)
        cell.alignment = C_ALN
        cell.border = BORDER

    for offset, tipo_key in enumerate(("DESARROLLO", "TAREA"), start=0):
        row = tipo_section_row + 2 + offset
        tipo_bg = "DBEAFE" if tipo_key == "DESARROLLO" else "FEF3C7"
        tipo_color = "1D4ED8" if tipo_key == "DESARROLLO" else "D97706"
        c1 = ws.cell(row=row, column=1, value=tipo_key)
        c2 = ws.cell(row=row, column=2, value=summary["tipo_counter"].get(tipo_key, 0))
        c3 = ws.cell(row=row, column=3, value=round(summary["tipo_horas"].get(tipo_key, 0), 2))
        for cell in (c1, c2, c3):
            cell.border = BORDER
            cell.fill = _fill(tipo_bg)
        c1.font = _font(bold=True, color=tipo_color)
        c1.alignment = L_ALN
        c2.font = _font(bold=True, color=C_BLUE)
        c2.alignment = C_ALN
        c3.font = _font(bold=True, color=C_BLUE)
        c3.alignment = C_ALN
        c3.number_format = "#,##0.0"


def _render_actividades(ws, proyecto_nombre, generado_en, actividades, evidencias_por_actividad, export_context=None):
    ws.title = "Actividades"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = "2563EB"
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:Q3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:3"

    ws.merge_cells("A1:Q1")
    ws["A1"] = proyecto_nombre
    ws["A1"].font = _font(bold=True, color=C_WHITE, size=16)
    ws["A1"].fill = _fill(C_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].border = BORDER
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:Q2")
    ws["A2"] = (
        f"Generado el {generado_en} · {_export_scope_short(export_context)} · "
        "La columna de evidencia resume tipos registrados, no archivos incrustados."
    )
    ws["A2"].font = _font(size=8, color="94A3B8")
    ws["A2"].alignment = C_ALN
    ws["A2"].fill = _fill(C_GRAY)
    ws["A2"].border = BORDER
    ws.row_dimensions[2].height = 18

    headers = [
        ("Nombre de la actividad", 36),
        ("Recursos utilizados", 22),
        ("Estatus", 16),
        ("Fecha solicitud", 14),
        ("Fecha inicio", 13),
        ("Fecha fin", 13),
        ("Prioridad", 11),
        ("Tipo", 14),
        ("Horas totales", 13),
        ("Responsable/s", 26),
        ("Solicitante", 18),
        ("Descripción", 34),
        ("Desglose", 44),
        ("Evidencia adjunta", 28),
        ("Actividades ligadas", 28),
        ("Observaciones", 34),
        ("% del total", 12),
    ]
    desglose_col_index = 13
    desglose_col_letter = get_column_letter(desglose_col_index)
    max_desglose_line_len = len("Desglose")
    total_horas_reporte = sum(float(a.get("HORAS_TOTALES") or 0) for a in actividades)

    ws.row_dimensions[3].height = 32
    for column_index, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=column_index, value=header)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_MID)
        cell.alignment = C_ALN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(column_index)].width = width

    std_row_height = 52
    for row_index, actividad in enumerate(actividades, start=4):
        actividad_id = actividad.get("ID")
        prioridad_texto = _prioridad(actividad.get("PRIORIDAD"))
        estatus_texto = _safe(actividad.get("ESTATUS"))
        horas = float(actividad.get("HORAS_TOTALES") or 0)
        row_bg = C_LBLUE if row_index % 2 == 0 else C_GRAY
        evidencia_texto = _tipos_evidencia_texto(evidencias_por_actividad.get(actividad_id, []))
        observaciones = actividad.get("NOTAS_REPORTE") or "—"
        tipo_actividad = _safe(actividad.get("TIPO"))
        desglose = _safe(actividad.get("DESGLOSE_REPORTE"))
        pct_del_total = horas / total_horas_reporte if total_horas_reporte > 0 else 0.0

        row_data = [
            _safe(actividad.get("NOMBRE_ACTIVIDAD")),
            _safe(actividad.get("RECURSOS")),
            estatus_texto,
            _fecha(actividad.get("FECHA_SOLICITUD")),
            _fecha(actividad.get("FECHA_INICIO")),
            _fecha(actividad.get("FECHA_FIN_REAL")),
            prioridad_texto,
            tipo_actividad,
            horas,
            _safe(actividad.get("RESPONSABLES")),
            _safe(actividad.get("SOLICITANTE")),
            _safe(actividad.get("DESCRIPCION")),
            desglose,
            evidencia_texto,
            _linked_text(actividad),
            observaciones,
            pct_del_total,
        ]

        desglose_lineas = str(desglose).splitlines() if desglose else ["—"]
        max_desglose_line_len = max(
            max_desglose_line_len,
            max((len(linea) for linea in desglose_lineas), default=0),
        )
        ws.row_dimensions[row_index].height = max(std_row_height, 16 * max(len(desglose_lineas), 1))
        for column_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.border = BORDER
            cell.font = _font()
            cell.fill = _fill(row_bg)

            if column_index in (4, 5, 6):
                cell.alignment = C_ALN
            elif column_index == 7:
                cell.alignment = C_ALN
                bg = {"Alta": C_ALTA, "Media": C_MEDIA, "Baja": C_BAJA}.get(prioridad_texto, C_WHITE)
                color = {"Alta": "DC2626", "Media": "D97706", "Baja": "16A34A"}.get(prioridad_texto, C_DARK)
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=color)
            elif column_index == 8:
                cell.alignment = C_ALN
                if tipo_actividad == "DESARROLLO":
                    cell.fill = _fill("DBEAFE")
                    cell.font = _font(bold=True, color="1D4ED8")
                elif tipo_actividad == "TAREA":
                    cell.fill = _fill("FEF3C7")
                    cell.font = _font(bold=True, color="D97706")
            elif column_index == 9:
                cell.alignment = C_ALN
                cell.number_format = "#,##0.0"
            elif column_index == 3:
                cell.alignment = C_ALN
                cell.fill = _fill(ESTATUS_BG.get(estatus_texto.lower(), C_WHITE))
            elif column_index == desglose_col_index:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif column_index == 17:
                cell.alignment = C_ALN
                cell.number_format = "0.0%"
            else:
                cell.alignment = L_ALN

        linked_cell = ws.cell(row=row_index, column=15)
        if actividad.get("ID_ACTIVIDAD_PADRE") and int(actividad.get("NUM_HIJAS") or 0) > 0:
            linked_cell.fill = _fill("E9D5FF")
        elif actividad.get("ID_ACTIVIDAD_PADRE"):
            linked_cell.fill = _fill(C_PARENT)
        elif int(actividad.get("NUM_HIJAS") or 0) > 0:
            linked_cell.fill = _fill(C_LINKED)
            linked_cell.font = _font(bold=True, color=C_BLUE)

        if horas > 40:
            horas_cell = ws.cell(row=row_index, column=9)
            horas_cell.fill = _fill("FEE2E2")
            horas_cell.font = _font(bold=True, color="DC2626")

        if int(actividad.get("ES_ACTIVIDAD_RAPIDA_HISTORICA") or 0) == 1:
            observation_cell = ws.cell(row=row_index, column=16)
            observation_cell.fill = _fill(C_WARN)
            observation_cell.border = WARN_BORDER_STYLE
            observation_cell.font = _font(color="9A3412", bold=True)
            ws.cell(row=row_index, column=1).font = _font(bold=True, color="9A3412")

    data_end_row = max(len(actividades) + 3, 3)
    ws.auto_filter.ref = f"A3:Q{data_end_row}"
    ws.column_dimensions[desglose_col_letter].width = max(34, min(70, max_desglose_line_len + 6))
    ws.print_area = f"A1:Q{data_end_row + 2}"

    total_row = len(actividades) + 4
    ws.row_dimensions[total_row].height = 20
    ws.merge_cells(f"A{total_row}:H{total_row}")

    total_caption = ws.cell(row=total_row, column=1, value="TOTAL HORAS")
    total_caption.font = _font(bold=True, color=C_WHITE)
    total_caption.fill = _fill(C_MID)
    total_caption.alignment = R_ALN
    total_caption.border = BORDER

    total_formula = f"=SUM(I4:I{total_row - 1})" if actividades else 0
    total_hours_cell = ws.cell(row=total_row, column=9, value=total_formula)
    total_hours_cell.font = _font(bold=True, color=C_WHITE)
    total_hours_cell.fill = _fill(C_BLUE)
    total_hours_cell.alignment = C_ALN
    total_hours_cell.number_format = "#,##0.0"
    total_hours_cell.border = BORDER

    total_pct_cell = ws.cell(row=total_row, column=17, value=1.0 if actividades else 0.0)
    total_pct_cell.font = _font(bold=True, color=C_WHITE)
    total_pct_cell.fill = _fill(C_BLUE)
    total_pct_cell.alignment = C_ALN
    total_pct_cell.number_format = "0.0%"
    total_pct_cell.border = BORDER

    for column_index in range(10, 17):
        cell = ws.cell(row=total_row, column=column_index)
        cell.fill = _fill(C_MID)
        cell.border = BORDER

    note_row = total_row + 2
    ws.merge_cells(f"A{note_row}:Q{note_row}")
    note = ws.cell(
        row=note_row,
        column=1,
        value=(
            "ℹ️ La columna DESGLOSE incluye los registros por actividad (desarrollador, acción/detalle y descripción). "
            "La columna de evidencia adjunta resume tipos de evidencia registrados. "
            "Las actividades rápidas históricas con datos parciales se marcan en Observaciones."
        ),
    )
    note.font = _font(size=8, color="94A3B8", italic=True)
    note.alignment = C_ALN


def _parse_desarrolladores(actividades):
    """Extrae stats de desarrolladores desde DESGLOSE_REPORTE."""
    devs = {}
    for actividad in actividades:
        act_id = actividad.get("ID") or actividad.get("NOMBRE_ACTIVIDAD")
        desglose = actividad.get("DESGLOSE_REPORTE") or ""
        current_dev = None
        dev_names_in_act = []
        for line in str(desglose).splitlines():
            if line and not line.startswith(" ") and line.rstrip().endswith(":"):
                dev_name = line.rstrip().rstrip(":")
                if dev_name not in devs:
                    devs[dev_name] = {"actividades": set(), "registros": 0}
                current_dev = dev_name
                if dev_name not in dev_names_in_act:
                    dev_names_in_act.append(dev_name)
                    devs[dev_name]["actividades"].add(act_id)
            elif line.strip().startswith("•") and current_dev:
                devs[current_dev]["registros"] += 1
    return {
        k: {"actividades": len(v["actividades"]), "registros": v["registros"]}
        for k, v in sorted(devs.items())
    }


def _render_desarrolladores(ws, actividades):
    """Hoja 3: tabla resumen por desarrollador extraída del DESGLOSE_REPORTE."""
    ws.title = "Por Desarrollador"
    ws.sheet_properties.tabColor = "16A34A"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "Resumen por desarrollador"
    title.font = _font(bold=True, color=C_WHITE, size=14)
    title.fill = _fill("16A34A")
    title.alignment = C_ALN
    title.border = BORDER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    subtitle = ws["A2"]
    subtitle.value = "Extraído del desglose por actividad. Los registros son entradas de trabajo individuales."
    subtitle.font = _font(size=8, color="94A3B8", italic=True)
    subtitle.alignment = C_ALN
    subtitle.fill = _fill(C_GRAY)
    subtitle.border = BORDER
    ws.row_dimensions[2].height = 16

    for col_i, label in enumerate(("Desarrollador", "Actividades", "Registros"), start=1):
        cell = ws.cell(row=3, column=col_i, value=label)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill("16A34A")
        cell.alignment = C_ALN
        cell.border = BORDER
    ws.row_dimensions[3].height = 24

    devs = _parse_desarrolladores(actividades)
    if not devs:
        empty = ws.cell(row=4, column=1, value="Sin datos de desglose disponibles.")
        empty.font = _font(italic=True, color="94A3B8")
        empty.alignment = L_ALN
        return

    for row_i, (dev, stats) in enumerate(devs.items(), start=4):
        row_bg = C_GRAY if row_i % 2 == 0 else C_WHITE
        c1 = ws.cell(row=row_i, column=1, value=dev)
        c2 = ws.cell(row=row_i, column=2, value=stats["actividades"])
        c3 = ws.cell(row=row_i, column=3, value=stats["registros"])
        for cell in (c1, c2, c3):
            cell.fill = _fill(row_bg)
            cell.border = BORDER
        c1.font = _font(bold=True, color=C_DARK)
        c1.alignment = L_ALN
        c2.font = _font(bold=True, color=C_BLUE)
        c2.alignment = C_ALN
        c3.font = _font(color=C_DARK)
        c3.alignment = C_ALN
        ws.row_dimensions[row_i].height = 18


def generar_reporte(
    proyecto_nombre: str,
    actividades: list,
    evidencias_por_actividad: dict,
    upload_base: str = "",
    export_context: dict | None = None,
) -> bytes:
    """
    Genera el Excel en memoria y devuelve bytes.

    La exportación incluye:
    - hoja de resumen ejecutivo
    - hoja de detalle cronológico
    - evidencia resumida por tipo
    - trazabilidad de actividades rápidas históricas incompletas
    """
    del upload_base  # Compatibilidad con la firma previa.

    wb = Workbook()
    generado_en = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws_resumen = wb.active
    _render_resumen(
        ws_resumen,
        proyecto_nombre,
        generado_en,
        actividades,
        evidencias_por_actividad,
        export_context=export_context,
    )

    ws_actividades = wb.create_sheet("Actividades")
    _render_actividades(
        ws_actividades,
        proyecto_nombre,
        generado_en,
        actividades,
        evidencias_por_actividad,
        export_context=export_context,
    )

    ws_desarrolladores = wb.create_sheet("Por Desarrollador")
    _render_desarrolladores(ws_desarrolladores, actividades)

    wb.active = 0

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
