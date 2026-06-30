"""
Generador de reporte Excel filtrado desde Registros.

- Hoja 1: Resumen ejecutivo
- Hoja 2: Detalle de actividades
- Hoja 3: Resumen de horas por desarrollador
- Evidencia resumida por tipo
"""

import io
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Nueva Paleta de Colores Moderna (Corporate Slate & Indigo) ──────────────
C_DARK = "0F172A"  # Slate 900 (Títulos principales y texto fuerte)
C_MID = "475569"  # Slate 600 (Encabezados secundarios)
C_BLUE = "1E40AF"  # Indigo/Blue (Énfasis institucional)
C_LBLUE = "F1F5F9"  # Slate 100 (Cebra en filas par)
C_GRAY = "F8FAFC"  # Slate 50 (Fondo alterno o sutil)
C_WHITE = "FFFFFF"

# Badges de Estatus y Prioridad (Pastel Soft)
C_ALTA = "FEE2E2"  # Soft Red
C_ALTA_T = "991B1B"
C_MEDIA = "FEF3C7"  # Soft Amber
C_MEDIA_T = "92400E"
C_BAJA = "DCFCE7"  # Soft Emerald
C_BAJA_T = "166534"

C_WARN = "FFEDD5"  # Soft Orange Alertas
C_WARN_BORDER = "FDBA74"
C_SUCCESS = "F0FDF4"  # Soft Green

ESTATUS_BG = {
    "en análisis": "F0F9FF",  # Soft Sky Blue
    "en proceso": "E0F2FE",  # Light Sky Blue
    "en espera de info del cliente": "FEF3C7",  # Amber
    "completado": "DCFCE7",  # Emerald
    "cancelado": "F1F5F9",  # Light Gray
}

ESTATUS_TXT = {
    "en análisis": "0369A1",
    "en proceso": "075985",
    "en espera de info del cliente": "92400E",
    "completado": "166534",
    "cancelado": "475569",
}

# Bordes más sutiles
THIN = Side(style="thin", color="E2E8F0")  # Slate 200
WARN_SIDE = Side(style="thin", color=C_WARN_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WARN_BORDER_STYLE = Border(
    left=WARN_SIDE, right=WARN_SIDE, top=WARN_SIDE, bottom=WARN_SIDE
)

C_ALN = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ALN = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ALN = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _fill(color):
    return PatternFill("solid", fgColor=color, start_color=color)


def _font(bold=False, color=C_DARK, size=9, italic=False):
    return Font(name="Segoe UI", bold=bold, color=color, size=size, italic=italic)


def _prioridad(value):
    return (
        {1: "Alta", 2: "Media", 3: "Baja"}.get(int(value), "—")
        if value is not None
        else "—"
    )


def _fecha(value):
    if value is None:
        return "—"
    text = str(value).strip()
    return text[:10] if len(text) >= 10 else text


def _safe(value):
    text = str(value).strip() if value is not None else ""
    return text or "—"


def _es_actividad_rapida(row):
    return str(row.get("NOMBRE_ACTIVIDAD") or "").startswith("⚡")


def _horas_directas(actividad):
    """
    Devuelve las horas que realmente le corresponden a esta actividad
    para efectos del total del reporte:

    - Padres con hijas: HORAS_DIRECTAS (solo registros no propagados, es decir
      los históricos propios del padre antes de que existieran hijas).
    - Hijas y actividades sueltas: HORAS_TOTALES (sus registros son siempre
      directos, nunca son espejos de nadie).

    De esta forma:
      · Las horas históricas del padre cuentan exactamente una vez.
      · Las horas de las hijas cuentan exactamente una vez.
      · Los espejos propagados del padre (copia de las hijas) nunca se suman.
    """
    es_padre_con_hijas = (
        not actividad.get("ID_ACTIVIDAD_PADRE")
        and int(actividad.get("NUM_HIJAS") or 0) > 0
    )
    if es_padre_con_hijas:
        return float(actividad.get("HORAS_DIRECTAS") or 0)
    return float(actividad.get("HORAS_TOTALES") or 0)


def _build_summary(actividades, evidencias_por_actividad):
    total_actividades = len(actividades)

    actividades_con_evidencia = sum(
        1 for item in actividades if item and evidencias_por_actividad.get(item.get("ID"))
    )
    actividades_rapidas = sum(1 for item in actividades if item and _es_actividad_rapida(item))
    actividades_rapidas_historicas = sum(
        1 for item in actividades if item and int(item.get("ES_ACTIVIDAD_RAPIDA_HISTORICA") or 0) == 1
    )

    estatus_counter = Counter()
    estatus_horas = Counter()
    prioridad_counter = Counter()
    tipo_counter = Counter()
    tipo_horas = Counter()

    total_horas = 0

    for item in actividades:
        if not item:
            continue

        estatus = _safe(item.get("ESTATUS"))
        tipo = _safe(item.get("TIPO"))
        prioridad = _prioridad(item.get("PRIORIDAD"))
        horas = _horas_directas(item)

        estatus_counter[estatus] += 1
        tipo_counter[tipo] += 1
        if prioridad != "—":
            prioridad_counter[prioridad] += 1

        total_horas += horas
        estatus_horas[estatus] += horas
        tipo_horas[tipo] += horas

    return {
        "total_actividades": total_actividades,
        "total_horas": total_horas,
        "actividades_con_evidencia": actividades_con_evidencia,
        "actividades_rapidas": actividades_rapidas,
        "actividades_rapidas_historicas": actividades_rapidas_historicas,
        "estatus_counter": estatus_counter,
        "estatus_horas": estatus_horas,
        "prioridad_counter": prioridad_counter,
        "tipo_counter": tipo_counter,
        "tipo_horas": tipo_horas,
    }


def _export_scope(export_context=None):
    export_context = export_context or {}
    return (
        export_context.get("scope_label")
        or "Todas las actividades del proyecto exportado"
    )


def _export_scope_short(export_context=None):
    export_context = export_context or {}
    return (
        export_context.get("scope_label_short") or "Detalle cronológico por actividad"
    )


def _render_resumen(
    ws,
    proyecto_nombre,
    generado_en,
    actividades,
    evidencias_por_actividad,
    export_context=None,
    total_row_actividades=None,
):
    summary = _build_summary(actividades, evidencias_por_actividad)

    ws.title = "Resumen"
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 100
    ws.sheet_properties.tabColor = C_DARK
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = f"  Resumen Ejecutivo · {proyecto_nombre}"
    title.font = _font(bold=True, color=C_WHITE, size=15)
    title.fill = _fill(C_BLUE)
    title.alignment = Alignment(horizontal="left", vertical="center")
    title.border = BORDER
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    subtitle = ws["A2"]
    subtitle.value = (
        "  Reporte orientado a cliente, PM y gerencia. Combina resumen ejecutivo, "
        "trazabilidad de exportación y detalle cronológico por actividad."
    )
    subtitle.font = _font(color=C_MID, size=9, italic=True)
    subtitle.fill = _fill(C_GRAY)
    subtitle.alignment = Alignment(horizontal="left", vertical="center")
    subtitle.border = BORDER
    ws.row_dimensions[2].height = 25

    ws.merge_cells("A4:F4")
    section = ws["A4"]
    section.value = " Contexto de exportación"
    section.font = _font(bold=True, color=C_WHITE, size=10)
    section.fill = _fill(C_MID)
    section.alignment = Alignment(horizontal="left", vertical="center")
    section.border = BORDER
    ws.row_dimensions[4].height = 22

    context_rows = [
        ("Proyecto", proyecto_nombre),
        ("Generado el", generado_en),
        ("Cobertura", _export_scope(export_context)),
        (
            "Criterio de evidencia",
            "La evidencia se resume por tipo; no se incrustan archivos",
        ),
    ]
    for index, (label, value) in enumerate(context_rows, start=5):
        left = ws.cell(row=index, column=1, value=f" {label}")
        right = ws.cell(row=index, column=2, value=value)
        for cell in (left, right):
            cell.border = BORDER
            cell.alignment = L_ALN
            cell.fill = _fill(C_WHITE if index % 2 else C_GRAY)
        left.font = _font(bold=True, color=C_MID)
        right.font = _font()
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=6)
        ws.row_dimensions[index].height = 20

    ws.merge_cells("A10:C10")
    summary_header = ws["A10"]
    summary_header.value = " Resumen general"
    summary_header.font = _font(bold=True, color=C_WHITE, size=10)
    summary_header.fill = _fill(C_MID)
    summary_header.alignment = Alignment(horizontal="left", vertical="center")
    summary_header.border = BORDER

    general_rows = [
        ("Actividades totales", summary["total_actividades"]),
        ("Horas totales", None),  # se escribe con fórmula abajo
        ("Actividades con evidencia", summary["actividades_con_evidencia"]),
        ("Actividades rápidas", summary["actividades_rapidas"]),
        (
            "Rápidas históricas con datos parciales",
            summary["actividades_rapidas_historicas"],
        ),
    ]
    for index, (label, value) in enumerate(general_rows, start=11):
        left = ws.cell(row=index, column=1, value=f" {label}")
        right = ws.cell(row=index, column=2, value=value)
        for cell in (left, right):
            cell.border = BORDER
            cell.alignment = C_ALN if cell.column == 2 else L_ALN
            cell.fill = _fill(C_WHITE if index % 2 else C_GRAY)
        left.font = _font(bold=True, color=C_DARK)
        right.font = _font(bold=True, color=C_BLUE, size=10)
        if label == "Horas totales":
            if total_row_actividades:
                right.value = (
                    f"=Actividades!I{total_row_actividades}"
                    f"+Actividades!J{total_row_actividades}"
                )
            right.number_format = "#,##0.0"
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=3)
        ws.row_dimensions[index].height = 20

    ws.merge_cells("D10:F10")
    status_header = ws["D10"]
    status_header.value = " Distribución por estatus"
    status_header.font = _font(bold=True, color=C_WHITE, size=10)
    status_header.fill = _fill(C_MID)
    status_header.alignment = Alignment(horizontal="left", vertical="center")
    status_header.border = BORDER

    ws["D11"] = "Estatus"
    ws["E11"] = "Actividades"
    ws["F11"] = "Horas"
    for cell in (ws["D11"], ws["E11"], ws["F11"]):
        cell.font = _font(bold=True, color=C_DARK)
        cell.fill = _fill(C_LBLUE)
        cell.alignment = C_ALN
        cell.border = BORDER
    ws.row_dimensions[11].height = 20

    status_start_row = 12
    for offset, estatus in enumerate(
        sorted(summary["estatus_counter"].keys()), start=0
    ):
        row = status_start_row + offset
        ws.cell(row=row, column=4, value=estatus)
        ws.cell(row=row, column=5, value=summary["estatus_counter"][estatus])
        ws.cell(row=row, column=6, value=round(summary["estatus_horas"][estatus], 2))

        estatus_norm = estatus.lower()
        bg_color = ESTATUS_BG.get(estatus_norm, C_WHITE)
        txt_color = ESTATUS_TXT.get(estatus_norm, C_DARK)

        for column in range(4, 7):
            cell = ws.cell(row=row, column=column)
            cell.border = BORDER
            if column == 4:
                cell.fill = _fill(bg_color)
                cell.font = _font(bold=True, color=txt_color)
                cell.alignment = L_ALN
            else:
                cell.fill = _fill(C_WHITE if row % 2 else C_GRAY)
                cell.font = _font(bold=True if column == 6 else False)
                cell.alignment = C_ALN

        ws.cell(row=row, column=6).number_format = "#,##0.0"
        ws.row_dimensions[row].height = 20

    prioridad_section_row = max(
        status_start_row + max(len(summary["estatus_counter"]), 1) + 2, 18
    )
    ws.merge_cells(
        start_row=prioridad_section_row,
        start_column=1,
        end_row=prioridad_section_row,
        end_column=2,
    )
    prioridad_header = ws.cell(
        row=prioridad_section_row, column=1, value=" Distribución por prioridad"
    )
    prioridad_header.font = _font(bold=True, color=C_WHITE, size=10)
    prioridad_header.fill = _fill(C_MID)
    prioridad_header.alignment = Alignment(horizontal="left", vertical="center")
    prioridad_header.border = BORDER
    ws.row_dimensions[prioridad_section_row].height = 22

    prioridad_labels = ("Prioridad", "Actividades")
    for offset, label in enumerate(prioridad_labels, start=0):
        cell = ws.cell(row=prioridad_section_row + 1, column=1 + offset, value=label)
        cell.font = _font(bold=True, color=C_DARK)
        cell.fill = _fill(C_LBLUE)
        cell.alignment = C_ALN
        cell.border = BORDER
    ws.row_dimensions[prioridad_section_row + 1].height = 20

    for offset, prioridad in enumerate(("Alta", "Media", "Baja"), start=0):
        row = prioridad_section_row + 2 + offset
        left = ws.cell(row=row, column=1, value=f" {prioridad}")
        right = ws.cell(
            row=row, column=2, value=summary["prioridad_counter"].get(prioridad, 0)
        )
        left.border = right.border = BORDER
        left.alignment = L_ALN
        right.alignment = C_ALN

        bg = {"Alta": C_ALTA, "Media": C_MEDIA, "Baja": C_BAJA}.get(prioridad, C_WHITE)
        color = {"Alta": C_ALTA_T, "Media": C_MEDIA_T, "Baja": C_BAJA_T}.get(
            prioridad, C_DARK
        )

        left.fill = _fill(bg)
        left.font = _font(bold=True, color=color)
        right.fill = _fill(C_WHITE if row % 2 else C_GRAY)
        right.font = _font(bold=True, color=C_BLUE)
        ws.row_dimensions[row].height = 20

    alert_row = prioridad_section_row + 6
    ws.merge_cells(start_row=alert_row, start_column=1, end_row=alert_row, end_column=6)
    alert_header = ws.cell(row=alert_row, column=1, value=" Alertas y notas")
    alert_header.font = _font(bold=True, color=C_WHITE, size=10)
    alert_header.fill = _fill(C_DARK)
    alert_header.alignment = Alignment(horizontal="left", vertical="center")
    alert_header.border = BORDER
    ws.row_dimensions[alert_row].height = 22

    alert_text = (
        "  Se detectaron actividades rápidas históricas con datos parciales. "
        "En la hoja 'Actividades' se marcan explícitamente para evitar interpretaciones incompletas."
        if summary["actividades_rapidas_historicas"] > 0
        else "  No se detectaron alertas de integridad en las actividades exportadas."
    )
    ws.merge_cells(
        start_row=alert_row + 1, start_column=1, end_row=alert_row + 2, end_column=6
    )
    alert_note = ws.cell(row=alert_row + 1, column=1, value=alert_text)
    alert_note.font = _font(
        color=C_DARK, italic=summary["actividades_rapidas_historicas"] == 0
    )
    alert_note.fill = _fill(
        C_WARN if summary["actividades_rapidas_historicas"] > 0 else C_SUCCESS
    )
    alert_note.border = (
        WARN_BORDER_STYLE if summary["actividades_rapidas_historicas"] > 0 else BORDER
    )
    alert_note.alignment = L_ALN

    tipo_section_row = alert_row + 4
    ws.merge_cells(
        start_row=tipo_section_row,
        start_column=1,
        end_row=tipo_section_row,
        end_column=3,
    )
    tipo_hdr = ws.cell(row=tipo_section_row, column=1, value=" Distribución por tipo")
    tipo_hdr.font = _font(bold=True, color=C_WHITE, size=10)
    tipo_hdr.fill = _fill(C_MID)
    tipo_hdr.alignment = Alignment(horizontal="left", vertical="center")
    tipo_hdr.border = BORDER
    ws.row_dimensions[tipo_section_row].height = 22

    for col_i, label in enumerate(("Tipo", "Actividades", "Horas"), start=1):
        cell = ws.cell(row=tipo_section_row + 1, column=col_i, value=label)
        cell.font = _font(bold=True, color=C_DARK)
        cell.fill = _fill(C_LBLUE)
        cell.alignment = C_ALN
        cell.border = BORDER
    ws.row_dimensions[tipo_section_row + 1].height = 20

    for offset, tipo_key in enumerate(("DESARROLLO", "TAREA"), start=0):
        row = tipo_section_row + 2 + offset
        tipo_bg = "EFF6FF" if tipo_key == "DESARROLLO" else "FEFBF0"
        tipo_color = "1E40AF" if tipo_key == "DESARROLLO" else "B45309"
        horas_col = "I" if tipo_key == "DESARROLLO" else "J"
        c1 = ws.cell(row=row, column=1, value=f" {tipo_key}")
        c2 = ws.cell(row=row, column=2, value=summary["tipo_counter"].get(tipo_key, 0))
        c3 = ws.cell(
            row=row,
            column=3,
            value=(
                f"=Actividades!{horas_col}{total_row_actividades}"
                if total_row_actividades
                else round(summary["tipo_horas"].get(tipo_key, 0), 2)
            ),
        )
        for cell in (c1, c2, c3):
            cell.border = BORDER
            cell.fill = _fill(tipo_bg)
        c1.font = _font(bold=True, color=tipo_color)
        c1.alignment = L_ALN
        c2.font = _font(bold=True, color=C_DARK)
        c2.alignment = C_ALN
        c3.font = _font(bold=True, color=C_DARK)
        c3.alignment = C_ALN
        c3.number_format = "#,##0.0"
        ws.row_dimensions[row].height = 20


def _render_actividades(ws, proyecto_nombre, generado_en, actividades, evidencias_por_actividad, export_context=None):
    ws.title = "Actividades"
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 100
    ws.sheet_properties.tabColor = C_BLUE
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:3"

    ws.merge_cells("A1:R1")
    ws["A1"] = f"  {proyecto_nombre}"
    ws["A1"].font = _font(bold=True, color=C_WHITE, size=15)
    ws["A1"].fill = _fill(C_BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].border = BORDER
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:R2")
    ws["A2"] = (
        f"  Generado el {generado_en} · {_export_scope_short(export_context)} · "
        "Estructura alineada con entregables de ambiente productivo."
    )
    ws["A2"].font = _font(size=8.5, color=C_MID, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].fill = _fill(C_GRAY)
    ws["A2"].border = BORDER
    ws.row_dimensions[2].height = 22

    headers = [
        ("Entregables Ambiente Productivo", 40),
        ("Descripción requerimiento", 34),
        ("Tipo de Desarrollo", 20),
        ("Desarrollo / Tarea", 15),
        ("Fecha de solicitud", 14),
        ("Fecha Inicio", 13),
        ("Fecha Fin", 13),
        ("Complejidad", 12),
        ("Horas Desarrollo", 15),
        ("Horas Tareas", 13),
        ("% Avance", 12),
        ("Estatus", 16),
        ("Responsable (Solicitante)", 26),
        ("Recurso", 22),
        ("Dependencia", 16),
        ("Actividades", 44),
        ("Comentarios", 20),
        # Columnas ocultas: valores numéricos reales para totales por tipo
        ("_horas_dev_real", 0),    # Columna R — desarrollo real (sin espejos)
        ("_horas_tarea_real", 0),  # Columna S — tareas real (sin espejos)
    ]

    actividades_col_index = 16
    actividades_col_letter = get_column_letter(actividades_col_index)
    horas_dev_real_col = 18   # Columna R oculta
    horas_tarea_real_col = 19  # Columna S oculta
    max_actividades_line_len = len("Actividades")

    ws.row_dimensions[3].height = 28
    for column_index, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=column_index, value=header)
        if header in ("_horas_dev_real", "_horas_tarea_real"):
            ws.column_dimensions[get_column_letter(column_index)].width = 0
            ws.column_dimensions[get_column_letter(column_index)].hidden = True
            cell.value = None
            continue
        cell.font = _font(bold=True, color=C_WHITE, size=9.5)
        cell.fill = _fill(C_DARK)
        cell.alignment = C_ALN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(column_index)].width = width

    # ── Reordenamiento Jerárquico ────────────────────────────────────────────
    lista_ordenada = []
    padres = [a for a in actividades if not a.get("ID_ACTIVIDAD_PADRE")]
    hijas = [a for a in actividades if a.get("ID_ACTIVIDAD_PADRE")]

    hijas_count_por_padre = {}
    for padre in padres:
        lista_ordenada.append(padre)
        padre_id = padre.get("ID")
        hijas_del_padre = [
            h for h in hijas if str(h.get("ID_ACTIVIDAD_PADRE")) == str(padre_id)
        ]
        hijas_count_por_padre[padre_id] = int(
            padre.get("NUM_HIJAS") or len(hijas_del_padre)
        )
        lista_ordenada.extend(hijas_del_padre)

    ids_en_lista = {id(a) for a in lista_ordenada}
    for act in actividades:
        if id(act) not in ids_en_lista:
            lista_ordenada.append(act)

    # ── Colores para agrupar visualmente cada bloque padre+hijas ────────────
    GROUP_COLORS = ["EFF6FF", "F0FDF4", "FEF9F0", "F5F3FF", "FDF2F8", "F0F9FF"]
    group_color_idx = -1
    current_group_color = None
    grupo_actual_padre_id = None

    std_row_height = 45
    for row_index, actividad in enumerate(lista_ordenada, start=4):
        prioridad_texto = _prioridad(actividad.get("PRIORIDAD"))
        estatus_texto = _safe(actividad.get("ESTATUS"))
        tipo_actividad = _safe(actividad.get("TIPO"))
        desglose_actividades = _safe(actividad.get("DESGLOSE_REPORTE"))

        es_hija = bool(actividad.get("ID_ACTIVIDAD_PADRE"))

        if not es_hija:
            group_color_idx = (group_color_idx + 1) % len(GROUP_COLORS)
            current_group_color = GROUP_COLORS[group_color_idx]
            grupo_actual_padre_id = actividad.get("ID")

        tiene_hijas = hijas_count_por_padre.get(grupo_actual_padre_id, 0) > 0
        row_bg = current_group_color if (tiene_hijas and current_group_color) else (
            C_WHITE if row_index % 2 == 0 else C_GRAY
        )

        # ── Nombre con jerarquía ─────────────────────────────────────────────
        nombre_original = _safe(actividad.get("NOMBRE_ACTIVIDAD"))
        if es_hija:
            nombre_reporte = f"      ↳ {nombre_original}"
        else:
            n_hijas = hijas_count_por_padre.get(actividad.get("ID"), 0)
            if n_hijas > 0:
                nombre_reporte = f"▾ {nombre_original}  ({n_hijas} subactividad{'es' if n_hijas != 1 else ''})"
            else:
                nombre_reporte = f"  {nombre_original}"

        try:
            avance_val = float(actividad.get("AVANCE_PCT") or 0)
            if avance_val > 1.0:
                avance_val = avance_val / 100.0
        except (ValueError, TypeError):
            avance_val = 0.0

        # ── Lógica de horas simplificada ─────────────────────────────────────
        # Cada actividad muestra y suma sus propias horas reales:
        #
        #   · Padre con hijas → HORAS_DIRECTAS (registros no propagados del padre,
        #     es decir sus horas históricas propias). Puede ser 0 si el padre
        #     nunca tuvo registros directos.
        #   · Hija o actividad suelta → HORAS_TOTALES (siempre son directas,
        #     nunca son espejos de nadie).
        #
        # Los espejos propagados del padre (copia de las hijas) nunca aparecen.
        # No hay herencia visual: cada fila muestra exactamente lo que le pertenece.
        horas_para_mostrar = _horas_directas(actividad)

        if tipo_actividad == "DESARROLLO":
            celda_horas_desarrollo = horas_para_mostrar if horas_para_mostrar else ""
            celda_horas_tareas = ""
            horas_dev_real = horas_para_mostrar if horas_para_mostrar else None
            horas_tarea_real = None
        elif tipo_actividad == "TAREA":
            celda_horas_desarrollo = ""
            celda_horas_tareas = horas_para_mostrar if horas_para_mostrar else ""
            horas_dev_real = None
            horas_tarea_real = horas_para_mostrar if horas_para_mostrar else None
        else:
            celda_horas_desarrollo = horas_para_mostrar if horas_para_mostrar else ""
            celda_horas_tareas = ""
            horas_dev_real = horas_para_mostrar if horas_para_mostrar else None
            horas_tarea_real = None

        dependencia_texto = (
            _safe(actividad.get("NOMBRE_ACTIVIDAD_PADRE")) if es_hija else "—"
        )

        row_data = [
            _safe(actividad.get("NOMBRE_ENTREGABLE")),   # A  col 1
            nombre_reporte,                               # B  col 2
            _safe(actividad.get("RECURSOS")),             # C  col 3
            tipo_actividad,                               # D  col 4
            _fecha(actividad.get("FECHA_SOLICITUD")),     # E  col 5
            _fecha(actividad.get("FECHA_INICIO")),        # F  col 6
            _fecha(actividad.get("FECHA_FIN_REAL")),      # G  col 7
            prioridad_texto,                              # H  col 8
            celda_horas_desarrollo,                       # I  col 9  (Horas Desarrollo)
            celda_horas_tareas,                           # J  col 10 (Horas Tareas)
            avance_val,                                   # K  col 11
            estatus_texto,                                # L  col 12
            _safe(actividad.get("SOLICITANTE")),          # M  col 13
            _safe(actividad.get("RESPONSABLES")),         # N  col 14
            dependencia_texto,                            # O  col 15
            desglose_actividades,                         # P  col 16
            "",                                           # Q  col 17 (Comentarios)
            horas_dev_real,                               # R  col 18 (oculta, dev real)
            horas_tarea_real,                             # S  col 19 (oculta, tarea real)
        ]

        desglose_lineas = (
            str(desglose_actividades).splitlines() if desglose_actividades else ["—"]
        )
        max_actividades_line_len = max(
            max_actividades_line_len,
            max((len(linea) for linea in desglose_lineas), default=0),
        )
        ws.row_dimensions[row_index].height = max(
            std_row_height, 15 * max(len(desglose_lineas), 1)
        )

        for column_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)

            # Columnas R y S ocultas: solo valor numérico, sin formato visible
            if column_index in (horas_dev_real_col, horas_tarea_real_col):
                if value is not None:
                    cell.number_format = "#,##0.0"
                continue

            cell.border = BORDER

            if column_index == 1:
                cell.alignment = C_ALN
                if es_hija:
                    cell.font = _font(bold=False, italic=True, color="555555")
                else:
                    cell.font = _font(bold=True, color=C_DARK)
            elif column_index == 2:
                cell.alignment = L_ALN
                if es_hija:
                    cell.font = _font(bold=False, italic=True, color="555555")
                else:
                    cell.font = _font(bold=True, color=C_DARK)
            else:
                cell.font = _font(bold=False, italic=es_hija, color="555555" if es_hija else C_DARK)

            cell.fill = _fill(row_bg)

            if column_index in (3, 5, 6, 7, 13, 14, 15):
                cell.alignment = C_ALN
            elif column_index == 4:
                cell.alignment = C_ALN
                if tipo_actividad == "DESARROLLO":
                    cell.fill = _fill("E0F2FE")
                    cell.font = _font(bold=True, color="0369A1")
                elif tipo_actividad == "TAREA":
                    cell.fill = _fill("FEF3C7")
                    cell.font = _font(bold=True, color="B45309")
            elif column_index == 8:
                cell.alignment = C_ALN
                bg = {"Alta": C_ALTA, "Media": C_MEDIA, "Baja": C_BAJA}.get(
                    prioridad_texto, C_WHITE
                )
                color = {"Alta": C_ALTA_T, "Media": C_MEDIA_T, "Baja": C_BAJA_T}.get(
                    prioridad_texto, C_DARK
                )
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=color)
            elif column_index in (9, 10):
                cell.alignment = C_ALN
                if value != "":
                    cell.number_format = "#,##0.0"
            elif column_index == 11:
                cell.alignment = C_ALN
                cell.number_format = "0.0%"
                if avance_val == 1.0:
                    cell.font = _font(bold=True, color="166534")
            elif column_index == 12:
                cell.alignment = C_ALN
                estatus_norm = estatus_texto.lower()
                cell.fill = _fill(ESTATUS_BG.get(estatus_norm, C_WHITE))
                cell.font = _font(
                    bold=True, color=ESTATUS_TXT.get(estatus_norm, C_DARK)
                )
            elif column_index == actividades_col_index:
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
            else:
                cell.alignment = L_ALN

        # Borde grueso a la izquierda en filas hijas
        if es_hija:
            existing = ws.cell(row=row_index, column=2).border
            thick_left = Side(style="medium", color=C_BLUE)
            ws.cell(row=row_index, column=2).border = Border(
                left=thick_left, right=existing.right, top=existing.top, bottom=existing.bottom
            )

        if int(actividad.get("ES_ACTIVIDAD_RAPIDA_HISTORICA") or 0) == 1:
            ws.cell(row=row_index, column=2).font = _font(bold=True, color="9A3412")

        if es_hija:
            ws.row_dimensions[row_index].outlineLevel = 1

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryAbove = True

    data_end_row = max(len(actividades) + 3, 3)
    ws.auto_filter.ref = f"A3:Q{data_end_row}"
    ws.column_dimensions[actividades_col_letter].width = max(
        34, min(70, max_actividades_line_len + 6)
    )
    ws.print_area = f"A1:Q{data_end_row + 2}"

    # ── Fila de totales ──────────────────────────────────────────────────────
    # Columna R (oculta) = horas DESARROLLO reales (sin espejos propagados)
    # Columna S (oculta) = horas TAREA reales (sin espejos propagados)
    total_row = len(actividades) + 4
    ws.row_dimensions[total_row].height = 24
    ws.merge_cells(f"A{total_row}:H{total_row}")

    total_caption = ws.cell(row=total_row, column=1, value="TOTAL HORAS ASIGNADAS   ")
    total_caption.font = _font(bold=True, color=C_WHITE, size=10)
    total_caption.fill = _fill(C_DARK)
    total_caption.alignment = R_ALN
    total_caption.border = BORDER

    formula_dev = f"=SUM(R4:R{total_row - 1})" if actividades else 0
    total_dev_cell = ws.cell(row=total_row, column=9, value=formula_dev)
    total_dev_cell.font = _font(bold=True, color=C_WHITE, size=10)
    total_dev_cell.fill = _fill(C_BLUE)
    total_dev_cell.alignment = C_ALN
    total_dev_cell.number_format = "#,##0.0"
    total_dev_cell.border = BORDER

    formula_tarea = f"=SUM(S4:S{total_row - 1})" if actividades else 0
    total_task_cell = ws.cell(row=total_row, column=10, value=formula_tarea)
    total_task_cell.font = _font(bold=True, color=C_WHITE, size=10)
    total_task_cell.fill = _fill(C_BLUE)
    total_task_cell.alignment = C_ALN
    total_task_cell.number_format = "#,##0.0"
    total_task_cell.border = BORDER

    for column_index in range(11, 20):
        cell = ws.cell(row=total_row, column=column_index)
        cell.fill = _fill(C_DARK)
        cell.border = BORDER

    # ── Nota al pie ──────────────────────────────────────────────────────────
    note_row = total_row + 2
    ws.merge_cells(f"A{note_row}:Q{note_row}")
    note = ws.cell(
        row=note_row,
        column=1,
        value=(
            " ℹ️ El presente reporte refleja la distribución de requerimientos, tipos de desarrollo, "
            "complejidades y horas asignadas con base al módulo de registros del sistema."
        ),
    )
    note.font = _font(size=8.5, color=C_MID, italic=True)
    note.alignment = L_ALN

    legend_row = note_row + 1
    ws.merge_cells(f"A{legend_row}:Q{legend_row}")
    legend = ws.cell(
        row=legend_row,
        column=1,
        value=(
            " ℹ️ Actividades padre: las horas mostradas son registros propios históricos (anteriores a la creación de subactividades). "
            "Las horas de las subactividades se muestran y contabilizan en sus propias filas."
        ),
    )
    legend.font = _font(size=8.5, color=C_MID, italic=True)
    legend.alignment = L_ALN
    legend.fill = _fill(C_GRAY)


def _render_desarrolladores(ws, actividades, resumen_desarrolladores=None):
    """
    Hoja "Por Desarrollador": dos secciones.

    1. Tabla resumen (vista rápida): por desarrollador, total de actividades,
       registros y horas (Desarrollo / Tarea / Total).
    2. Detalle por actividad: un bloque por desarrollador con la lista de
       actividades a las que corresponden esas horas, para poder auditar
       de dónde sale cada cifra del resumen.

    `resumen_desarrolladores` viene de obtener_datos_reporte_proyecto() y tiene
    la forma:
        {
            "Nombre Completo": {
                "horas_total": float,
                "horas_desarrollo": float,
                "horas_tarea": float,
                "registros": int,
                "actividades": int,
                "detalle": [
                    {"nombre": str, "tipo": "DESARROLLO"|"TAREA", "horas": float, "registros": int},
                    ...
                ],
            },
            ...
        }
    """
    ws.title = "Por Desarrollador"
    ws.sheet_properties.tabColor = "10B981"
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 100
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 42
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 17

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "  Resumen por desarrollador"
    title.font = _font(bold=True, color=C_WHITE, size=14)
    title.fill = _fill("047857")
    title.alignment = Alignment(horizontal="left", vertical="center")
    title.border = BORDER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    subtitle = ws["A2"]
    subtitle.value = "  Horas reales registradas por cada desarrollador en el período exportado."
    subtitle.font = _font(size=8.5, color=C_MID, italic=True)
    subtitle.alignment = Alignment(horizontal="left", vertical="center")
    subtitle.fill = _fill(C_GRAY)
    subtitle.border = BORDER
    ws.row_dimensions[2].height = 20

    headers = ("Desarrollador", "Actividades", "Registros", "Horas Desarrollo", "Horas Tareas", "Horas Totales")
    for col_i, label in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_i, value=label)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_DARK)
        cell.alignment = C_ALN
        cell.border = BORDER
    ws.row_dimensions[3].height = 24

    devs = resumen_desarrolladores or {}
    if not devs:
        ws.merge_cells("A4:F4")
        empty = ws.cell(row=4, column=1, value="Sin datos de horas disponibles para este filtro.")
        empty.font = _font(italic=True, color=C_MID)
        empty.alignment = L_ALN
        empty.border = BORDER
        return

    # Orden de mayor a menor aporte de horas totales (se reutiliza en ambas secciones)
    devs_ordenados = sorted(devs.items(), key=lambda kv: kv[1].get("horas_total", 0), reverse=True)

    # ── Sección 1: tabla resumen ─────────────────────────────────────────────
    row_i = 4
    for dev, stats in devs_ordenados:
        row_bg = C_WHITE if row_i % 2 == 0 else C_GRAY
        valores = [
            f" {dev}",
            stats.get("actividades", 0),
            stats.get("registros", 0),
            round(stats.get("horas_desarrollo", 0), 2),
            round(stats.get("horas_tarea", 0), 2),
            round(stats.get("horas_total", 0), 2),
        ]
        for col_i, value in enumerate(valores, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.fill = _fill(row_bg)
            cell.border = BORDER
            if col_i == 1:
                cell.font = _font(bold=True, color=C_DARK)
                cell.alignment = L_ALN
            elif col_i in (4, 5, 6):
                cell.font = _font(bold=(col_i == 6), color=C_BLUE if col_i == 6 else C_DARK)
                cell.alignment = C_ALN
                cell.number_format = "#,##0.0"
            else:
                cell.font = _font(color=C_DARK)
                cell.alignment = C_ALN
        ws.row_dimensions[row_i].height = 20
        row_i += 1

    total_row = row_i
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    label_cell = ws.cell(row=total_row, column=1, value=" TOTAL")
    label_cell.font = _font(bold=True, color=C_WHITE)
    label_cell.fill = _fill(C_DARK)
    label_cell.alignment = R_ALN
    label_cell.border = BORDER

    total_dev = sum(s.get("horas_desarrollo", 0) for s in devs.values())
    total_tarea = sum(s.get("horas_tarea", 0) for s in devs.values())
    total_gen = sum(s.get("horas_total", 0) for s in devs.values())
    for col_i, value in ((4, total_dev), (5, total_tarea), (6, total_gen)):
        cell = ws.cell(row=total_row, column=col_i, value=round(value, 2))
        cell.number_format = "#,##0.0"
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_DARK)
        cell.alignment = C_ALN
        cell.border = BORDER
    ws.row_dimensions[total_row].height = 22

    # ── Sección 2: detalle por actividad dentro de cada desarrollador ───────
    detail_header_row = total_row + 3
    ws.merge_cells(start_row=detail_header_row, start_column=1, end_row=detail_header_row, end_column=6)
    detail_header = ws.cell(
        row=detail_header_row, column=1,
        value=" Detalle de horas por actividad y desarrollador",
    )
    detail_header.font = _font(bold=True, color=C_WHITE, size=10)
    detail_header.fill = _fill(C_DARK)
    detail_header.alignment = Alignment(horizontal="left", vertical="center")
    detail_header.border = BORDER
    ws.row_dimensions[detail_header_row].height = 22

    detail_sub = detail_header_row + 1
    ws.merge_cells(start_row=detail_sub, start_column=1, end_row=detail_sub, end_column=6)
    detail_sub_cell = ws.cell(
        row=detail_sub, column=1,
        value="  Para cada desarrollador, las actividades a las que corresponden sus horas (ordenadas de mayor a menor aporte).",
    )
    detail_sub_cell.font = _font(size=8.5, color=C_MID, italic=True)
    detail_sub_cell.fill = _fill(C_GRAY)
    detail_sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    detail_sub_cell.border = BORDER
    ws.row_dimensions[detail_sub].height = 20

    row_cursor = detail_sub + 2

    for dev, stats in devs_ordenados:
        # Banda con el nombre del desarrollador y sus totales
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=6)
        resumen_txt = (
            f"  {dev}   ·   {stats.get('actividades', 0)} actividad(es)   ·   "
            f"{stats.get('registros', 0)} registro(s)   ·   "
            f"{round(stats.get('horas_total', 0), 1):g}h totales "
            f"(Desarrollo: {round(stats.get('horas_desarrollo', 0), 1):g}h · "
            f"Tareas: {round(stats.get('horas_tarea', 0), 1):g}h)"
        )
        dev_band = ws.cell(row=row_cursor, column=1, value=resumen_txt)
        dev_band.font = _font(bold=True, color=C_WHITE, size=9.5)
        dev_band.fill = _fill("059669")
        dev_band.alignment = Alignment(horizontal="left", vertical="center")
        dev_band.border = BORDER
        ws.row_dimensions[row_cursor].height = 22
        row_cursor += 1

        # Encabezado de columnas del bloque de detalle
        sub_headers = ("Actividad", "Tipo", "Horas", "Registros")
        for col_i, label in enumerate(sub_headers, start=1):
            cell = ws.cell(row=row_cursor, column=col_i, value=label)
            cell.font = _font(bold=True, color=C_DARK, size=8.5)
            cell.fill = _fill(C_LBLUE)
            cell.alignment = L_ALN if col_i == 1 else C_ALN
            cell.border = BORDER
        # Columnas E y F se mantienen vacías pero con borde para que el bloque se vea uniforme
        for col_i in (5, 6):
            cell = ws.cell(row=row_cursor, column=col_i, value="")
            cell.fill = _fill(C_LBLUE)
            cell.border = BORDER
        ws.row_dimensions[row_cursor].height = 18
        row_cursor += 1

        detalle = stats.get("detalle") or []
        if not detalle:
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=4)
            empty_cell = ws.cell(
                row=row_cursor, column=1,
                value="  Sin actividades formales asociadas (horas registradas directamente).",
            )
            empty_cell.font = _font(italic=True, color=C_MID, size=8.5)
            empty_cell.alignment = L_ALN
            empty_cell.border = BORDER
            ws.row_dimensions[row_cursor].height = 18
            row_cursor += 1
        else:
            for idx, item in enumerate(detalle):
                row_bg = C_WHITE if idx % 2 == 0 else C_GRAY
                tipo_item = item.get("tipo") or "—"
                tipo_bg = "E0F2FE" if tipo_item == "DESARROLLO" else "FEF3C7"
                tipo_color = "0369A1" if tipo_item == "DESARROLLO" else "B45309"

                c1 = ws.cell(row=row_cursor, column=1, value=f"  ↳ {item.get('nombre', '—')}")
                c2 = ws.cell(row=row_cursor, column=2, value=tipo_item)
                c3 = ws.cell(row=row_cursor, column=3, value=round(item.get("horas", 0), 2))
                c4 = ws.cell(row=row_cursor, column=4, value=item.get("registros", 0))

                c1.font = _font(color=C_DARK, size=8.5)
                c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c1.fill = _fill(row_bg)

                c2.font = _font(bold=True, color=tipo_color, size=8.5)
                c2.alignment = C_ALN
                c2.fill = _fill(tipo_bg)

                c3.font = _font(bold=True, color=C_DARK, size=8.5)
                c3.alignment = C_ALN
                c3.fill = _fill(row_bg)
                c3.number_format = "#,##0.0"

                c4.font = _font(color=C_DARK, size=8.5)
                c4.alignment = C_ALN
                c4.fill = _fill(row_bg)

                for col_i in (5, 6):
                    cell = ws.cell(row=row_cursor, column=col_i, value="")
                    cell.fill = _fill(row_bg)

                for col_i in range(1, 7):
                    ws.cell(row=row_cursor, column=col_i).border = BORDER

                ws.row_dimensions[row_cursor].height = 20
                row_cursor += 1

        # Fila espaciadora entre desarrolladores
        row_cursor += 1



def generar_reporte(
    proyecto_nombre: str,
    actividades: list,
    evidencias_por_actividad: dict,
    upload_base: str = "",
    export_context: dict | None = None,
    resumen_desarrolladores: dict | None = None,
) -> bytes:
    del upload_base

    wb = Workbook()
    generado_en = datetime.now().strftime("%d/%m/%Y %H:%M")

    total_row_actividades = len(actividades) + 4

    ws_resumen = wb.active
    _render_resumen(
        ws_resumen,
        proyecto_nombre,
        generado_en,
        actividades,
        evidencias_por_actividad,
        export_context=export_context,
        total_row_actividades=total_row_actividades,
    )

    ws_actividades = wb.create_sheet("Actividades")
    _render_actividades(
        ws_actividades,
        proyecto_nombre,
        generado_en,
        actividades,
        evidencias_por_actividad,
        export_context=export_context,
    )

    ws_desarrolladores = wb.create_sheet("Por Desarrollador")
    _render_desarrolladores(ws_desarrolladores, actividades, resumen_desarrolladores)

    wb.active = 0

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# Alias por compatibilidad: routes.py importa "generar_reporte_proyecto_xlsx".
# Si tu routes.py real usa otro nombre, ajusta el import o este alias en consecuencia.
generar_reporte_proyecto_xlsx = generar_reporte